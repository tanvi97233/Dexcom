import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from discovery import ChromeSearchProvider, BrowserSettings, discovery_queries
from tracker_engine import JobResult, SHEET_NAME, canonical_url, country_from_location, is_dexcom_employer, job_id_from_url, parse_posted_datetime, update_tracker, validate, within_window
from dexcom_tracker import copy_to_downloads

class TrackerEngineTests(unittest.TestCase):
    def setUp(self): self.now = datetime(2026, 9, 3, 9, 0)
    def test_precise_time_windows(self):
        for value, expected in [("23 hours ago", True), ("24 hours ago", True), ("25 hours ago", False)]:
            self.assertEqual(within_window(parse_posted_datetime(value, self.now), self.now, "24hours"), expected)
        for value, expected in [("6 days ago", True), ("7 days ago", True), ("8 days ago", False)]:
            self.assertEqual(within_window(parse_posted_datetime(value, self.now), self.now, "7days"), expected)
    def test_countries_and_unknown(self):
        self.assertEqual(country_from_location("London, UK"), "United Kingdom")
        self.assertEqual(country_from_location("Austin, USA"), "United States")
        self.assertEqual(country_from_location("Seattle, US"), "United States")
        self.assertEqual(country_from_location("Seoul, Republic of Korea"), "South Korea")
        self.assertEqual(country_from_location("San Diego, CA"), "United States")
        self.assertEqual(country_from_location("Unresolved Place"), "")
    def test_canonical_url_variants(self):
        urls = ["https://linkedin.com/jobs/view/123/", "https://www.linkedin.com/jobs/view/123/?trk=x", "https://in.linkedin.com/jobs/view/sample-at-dexcom-123"]
        self.assertEqual({job_id_from_url(url) for url in urls}, {"123"})
        self.assertEqual(canonical_url(urls[2])[0], "https://www.linkedin.com/jobs/view/123/")
    def test_employer_validation(self):
        good = JobResult("https://www.linkedin.com/jobs/view/1/", "Dexcom hiring Engineer in Dublin, Ireland", "", "Dublin, Ireland", "today")
        bad = JobResult("https://www.linkedin.com/jobs/view/2/", "Engineer", "Other Company", "Dublin, Ireland", "today", "Working with Dexcom technology")
        self.assertTrue(is_dexcom_employer(good)); self.assertFalse(is_dexcom_employer(bad))
    def test_metadata_mapping_and_generic_queries(self):
        item = ChromeSearchProvider._metadata("Dexcom hiring Business Development Lead in Itanagar, Arunachal Pradesh, India | LinkedIn", "Dexcom Itanagar, Arunachal Pradesh, India. 2 hours ago", "https://in.linkedin.com/jobs/view/role-at-dexcom-4456463987")
        self.assertEqual((item.title, item.location, item.posted_date), ("Business Development Lead", "Itanagar, Arunachal Pradesh, India", "2 hours ago"))
        self.assertGreaterEqual(len(discovery_queries()), 4)
    def test_excel_append_and_deduplication(self):
        raw = [JobResult("https://www.linkedin.com/jobs/view/500/?x=y", "Dexcom A", "Dexcom", "Dublin, Ireland", "2 days ago"), JobResult("https://linkedin.com/jobs/view/500/", "Dexcom A", "Dexcom", "Ireland", "2 days ago"), JobResult("https://www.linkedin.com/jobs/view/501/", "Dexcom B", "Dexcom", "India", "today")]
        records = [validate(x, self.now)[0] for x in raw]
        with tempfile.TemporaryDirectory() as d:
            path=Path(d)/"tracker.xlsx"; self.assertEqual(update_tracker(records, path), (2,1)); self.assertEqual(update_tracker(records, path), (0,3)); ws=load_workbook(path)[SHEET_NAME]
            self.assertEqual(ws.max_row, 3); self.assertTrue(ws.cell(2,4).hyperlink)

    def test_completed_workbook_is_copied_to_downloads(self):
        with tempfile.TemporaryDirectory() as d:
            root = Path(d)
            workbook = root / "Dexcom_Job_Tracker.xlsx"
            workbook.write_bytes(b"new workbook")
            downloads = root / "missing-downloads"
            destination = copy_to_downloads(workbook, downloads)
            self.assertEqual(destination, downloads / workbook.name)
            self.assertEqual(destination.read_bytes(), b"new workbook")
            workbook.write_bytes(b"replacement workbook")
            copy_to_downloads(workbook, downloads)
            self.assertEqual(destination.read_bytes(), b"replacement workbook")

if __name__ == "__main__": unittest.main()
