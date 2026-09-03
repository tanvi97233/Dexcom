"""Provider-independent normalization, validation, deduplication, and Excel writing."""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["Country", "Job Posted Date", "Job Title", "Direct URL"]
SHEET_NAME = "Dexcom Job Tracker"
COUNTRY_NAMES = set("""Afghanistan|Albania|Algeria|Andorra|Angola|Argentina|Armenia|Australia|Austria|Azerbaijan|Bahamas|Bahrain|Bangladesh|Barbados|Belarus|Belgium|Belize|Benin|Bhutan|Bolivia|Bosnia and Herzegovina|Botswana|Brazil|Brunei|Bulgaria|Cambodia|Cameroon|Canada|Chile|China|Colombia|Costa Rica|Croatia|Cuba|Cyprus|Czechia|Denmark|Ecuador|Egypt|El Salvador|Estonia|Ethiopia|Finland|France|Georgia|Germany|Ghana|Greece|Guatemala|Honduras|Hong Kong|Hungary|Iceland|India|Indonesia|Iran|Iraq|Ireland|Israel|Italy|Jamaica|Japan|Jordan|Kazakhstan|Kenya|Kuwait|Latvia|Lebanon|Libya|Liechtenstein|Lithuania|Luxembourg|Malaysia|Malta|Mauritius|Mexico|Moldova|Monaco|Mongolia|Montenegro|Morocco|Myanmar|Namibia|Nepal|Netherlands|New Zealand|Nicaragua|Nigeria|North Korea|North Macedonia|Norway|Oman|Pakistan|Panama|Paraguay|Peru|Philippines|Poland|Portugal|Qatar|Romania|Russia|Rwanda|Saudi Arabia|Senegal|Serbia|Singapore|Slovakia|Slovenia|South Africa|South Korea|Spain|Sri Lanka|Sweden|Switzerland|Taiwan|Tajikistan|Tanzania|Thailand|Tunisia|Turkey|Uganda|Ukraine|United Arab Emirates|United Kingdom|United States|Uruguay|Uzbekistan|Venezuela|Vietnam|Yemen|Zambia|Zimbabwe|Republic of Korea""".split("|"))
# Multi-word countries are matched separately; aliases cover common public-search variants.
COUNTRY_NAMES.update({"United States", "United Kingdom", "United Arab Emirates", "South Korea", "Republic of Korea", "New Zealand", "Costa Rica", "South Africa", "Saudi Arabia", "North Macedonia", "Bosnia and Herzegovina", "Dominican Republic", "El Salvador", "Papua New Guinea", "Trinidad and Tobago", "Czech Republic"})
COUNTRY_ALIASES = {"uk":"United Kingdom", "u.k.":"United Kingdom", "usa":"United States", "u.s.a.":"United States", "us":"United States", "u.s.":"United States", "uae":"United Arab Emirates", "republic of korea":"South Korea", "south korea":"South Korea", "korea":"South Korea", "czechia":"Czech Republic"}
US_STATE_CODES = set("AL AK AZ AR CA CO CT DE FL GA HI ID IL IN IA KS KY LA ME MD MA MI MN MS MO MT NE NV NH NJ NM NY NC ND OH OK OR PA RI SC SD TN TX UT VT VA WA WV WI WY DC".split())

@dataclass(frozen=True)
class JobResult:
    direct_url: str
    title: str = ""
    company: str = ""
    location: str = ""
    posted_date: str = ""
    snippet: str = ""

@dataclass(frozen=True)
class JobRecord:
    country: str
    posted_at: datetime
    title: str
    url: str
    job_id: str

def clean(value: Any) -> str:
    return str(value or "").strip()

def job_id_from_url(url: str) -> str | None:
    parsed = urlparse(clean(url))
    match = re.search(r"/jobs/view/(\d+)(?:/|$)", parsed.path, re.I)
    if match:
        return match.group(1)
    match = re.search(r"-(\d+)(?:/|$)", parsed.path)
    if match:
        return match.group(1)
    values = parse_qs(parsed.query).get("currentJobId", [])
    return values[0] if values and values[0].isdigit() else None

def canonical_url(url: str) -> tuple[str | None, str | None]:
    parsed = urlparse(clean(url))
    host = parsed.netloc.casefold().removeprefix("www.")
    if parsed.scheme not in {"http", "https"} or not (host == "linkedin.com" or host.endswith(".linkedin.com")):
        return None, None
    job_id = job_id_from_url(url)
    return (f"https://www.linkedin.com/jobs/view/{job_id}/", job_id) if job_id else (None, None)

def country_from_location(location: str) -> str:
    raw = clean(location)
    parts = [part.strip() for part in raw.split(",") if part.strip()]
    for part in reversed(parts):
        normalized = part.casefold().rstrip(".")
        if normalized in COUNTRY_ALIASES: return COUNTRY_ALIASES[normalized]
        for country in COUNTRY_NAMES:
            if normalized == country.casefold(): return country
    if parts and parts[-1].upper().rstrip(".") in US_STATE_CODES:
        return "United States"
    return ""

def parse_posted_datetime(value: str, reference: datetime) -> datetime | None:
    value = clean(value).casefold()
    if value in {"just now", "today"}:
        return reference
    match = re.fullmatch(r"(\d+)\s+minute(?:s)?\s+ago", value)
    if match: return reference - timedelta(minutes=int(match.group(1)))
    match = re.fullmatch(r"(\d+)\s+hour(?:s)?\s+ago", value)
    if match: return reference - timedelta(hours=int(match.group(1)))
    match = re.fullmatch(r"(\d+)\s+day(?:s)?\s+ago", value)
    if match:
        return reference - timedelta(days=int(match.group(1)))
    match = re.fullmatch(r"(\d+)\s+week(?:s)?\s+ago", value)
    if match:
        return reference - timedelta(days=7 * int(match.group(1)))
    for pattern in ("%Y-%m-%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            # Explicit dates have no time source; treat them as local midnight consistently.
            return datetime.combine(datetime.strptime(value, pattern).date(), time.min)
        except ValueError:
            pass
    return None

def parse_posted_date(value: str, reference: datetime) -> date | None:
    parsed = parse_posted_datetime(value, reference)
    return parsed.date() if parsed else None

def is_dexcom_employer(result: JobResult) -> bool:
    company = clean(result.company).casefold()
    title = clean(result.title)
    snippet = clean(result.snippet)
    if company and (company == "dexcom" or company.startswith("dexcom ") or company.startswith("dexcom-")):
        return True
    return bool(re.match(r"^dexcom(?:\s+[^|]+)?\s+hiring\s+", title, re.I) or
                re.search(r"\bat\s+dexcom(?:\b|$)", title, re.I) or
                re.match(r"^dexcom(?:\s+[^,.]+)?\s+[A-Z][^,]+,", snippet, re.I))

def validate(result: JobResult, reference: datetime) -> tuple[JobRecord | None, str | None]:
    if not is_dexcom_employer(result):
        return None, "employer could not be validated as Dexcom"
    url, job_id = canonical_url(result.direct_url)
    if not url:
        return None, "invalid LinkedIn job URL"
    country = country_from_location(result.location)
    if not country:
        return None, "country/location unavailable"
    posted = parse_posted_datetime(result.posted_date, reference)
    if not posted:
        return None, "posting date unavailable or invalid"
    title = clean(result.title)
    if not title:
        return None, "job title unavailable"
    return JobRecord(country, posted, title, url, job_id), None

def within_window(posted: datetime, reference: datetime, filter_name: str) -> bool:
    return posted >= reference - timedelta(days=7 if filter_name == "7days" else 1)

def load_tracker(path: Path):
    if path.exists():
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Existing workbook is missing '{SHEET_NAME}'.")
        return wb, wb[SHEET_NAME]
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for i, header in enumerate(HEADERS, 1):
        c = ws.cell(1, i, header)
        c.font, c.fill = Font(name="Arial", bold=True, color="FFFFFF"), PatternFill("solid", fgColor="1F4E78")
    return wb, ws

def existing_ids(ws) -> set[str]:
    return {job_id for (url,) in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) if (job_id := job_id_from_url(clean(url)))}

def update_tracker(records: list[JobRecord], output: Path) -> tuple[int, int]:
    wb, ws = load_tracker(output)
    known = existing_ids(ws)
    added = skipped = 0
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for record in records:
        if record.job_id in known:
            skipped += 1
            continue
        known.add(record.job_id)
        row = ws.max_row + 1
        values = [record.country, record.posted_at.date(), record.title, record.url]
        for col, value in enumerate(values, 1):
            c = ws.cell(row, col, value)
            c.font, c.border = Font(name="Arial", size=10), border
            c.alignment = Alignment(vertical="top", wrap_text=(col == 3))
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
        ws.cell(row, 4).hyperlink = record.url
        ws.cell(row, 4).font = Font(name="Arial", size=10, color="0563C1", underline="single")
        added += 1
    ws.freeze_panes, ws.auto_filter.ref = "A2", f"A1:D{max(ws.max_row, 1)}"
    for i, width in enumerate([22, 18, 55, 60], 1): ws.column_dimensions[chr(64+i)].width = width
    wb.save(output)
    return added, skipped
