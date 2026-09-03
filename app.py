#!/usr/bin/env python3
"""Local web dashboard for the existing Dexcom LinkedIn job tracker.

This module deliberately does not contain any discovery, validation, date-window,
deduplication, or workbook-writing rules.  A run launches the established CLI and
the API reads its resulting workbook for the dashboard.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import threading
import uuid
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

from tracker_engine import HEADERS, SHEET_NAME

ROOT = Path(__file__).resolve().parent
WORKBOOK = Path(os.getenv("OUTPUT_FILE", ROOT / "Dexcom_Job_Tracker.xlsx")).resolve()
STATIC = ROOT / "web"
STATS_LABELS = {
    "job_cards_discovered": "LinkedIn job cards discovered",
    "valid_linkedin_urls": "Valid LinkedIn job URLs",
    "unique_jobs": "Unique jobs",
    "dexcom_jobs": "Dexcom jobs",
    "jobs_within_date_range": "Jobs within date range",
    "validation_failures": "Validation failures",
    "duplicates_removed": "Duplicates removed",
    "existing_jobs_skipped": "Existing jobs skipped",
    "new_jobs_added": "New jobs added",
}

state_lock = threading.Lock()
state: dict = {"status": "ready", "last_run": None, "runs": {}}


def read_jobs() -> list[dict]:
    """Read only the workbook produced by tracker_engine.update_tracker."""
    if not WORKBOOK.exists():
        return []
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            return []
        sheet = workbook[SHEET_NAME]
        if [sheet.cell(1, column=index).value for index in range(1, 5)] != HEADERS:
            raise ValueError("The tracker workbook has unexpected columns.")
        jobs = []
        for country, posted_date, title, url in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            if not url:
                continue
            jobs.append({
                "country": str(country or ""),
                "posted_date": posted_date.isoformat() if hasattr(posted_date, "isoformat") else str(posted_date or ""),
                "title": str(title or ""),
                "url": str(url),
            })
        return jobs
    finally:
        workbook.close()


def parse_stats(output: str) -> dict:
    stats = {}
    for key, label in STATS_LABELS.items():
        match = re.search(rf"^{re.escape(label)}:\s*(\d+)\s*$", output, re.MULTILINE)
        stats[key] = int(match.group(1)) if match else 0
    return stats


def launch_run(run_id: str, filter_name: str) -> None:
    command = [sys.executable, "dexcom_tracker.py", "--filter", filter_name, "--verbose"]
    environment = os.environ.copy()
    environment["PYTHONPATH"] = str(ROOT) + os.pathsep + environment.get("PYTHONPATH", "")
    try:
        result = subprocess.run(command, cwd=ROOT, env=environment, text=True, capture_output=True)
        output = (result.stdout + "\n" + result.stderr).strip()
        with state_lock:
            run = state["runs"][run_id]
            run.update({
                "status": "completed" if result.returncode == 0 else "failed",
                "completed_at": datetime.now().isoformat(timespec="seconds"),
                "stats": parse_stats(output) if result.returncode == 0 else None,
                "error": None if result.returncode == 0 else "The tracker process did not complete successfully.",
                "details": output[-12000:],
            })
            state["status"] = run["status"]
            state["last_run"] = run
    except Exception as error:  # defensive: the API itself should stay available
        with state_lock:
            run = state["runs"][run_id]
            run.update({"status": "failed", "completed_at": datetime.now().isoformat(timespec="seconds"), "error": str(error), "details": repr(error)})
            state["status"] = "failed"
            state["last_run"] = run


class DashboardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC), **kwargs)

    def log_message(self, format, *args):
        print("[dashboard] " + format % args)

    def send_json(self, payload: dict, status=HTTPStatus.OK):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/status":
            with state_lock:
                last_run = state["last_run"]
                payload = {"status": state["status"], "last_run": last_run}
            return self.send_json(payload)
        if parsed.path.startswith("/api/run/"):
            run_id = parsed.path.rsplit("/", 1)[-1]
            with state_lock:
                run = state["runs"].get(run_id)
            return self.send_json(run or {"error": "Run not found."}, HTTPStatus.OK if run else HTTPStatus.NOT_FOUND)
        if parsed.path == "/api/jobs":
            try:
                jobs = read_jobs()
                filter_name = parse_qs(parsed.query).get("filter", [None])[0]
                # The workbook is the source of persisted jobs; filter is retained as API-compatible metadata.
                return self.send_json({"jobs": jobs, "filter": filter_name})
            except Exception as error:
                return self.send_json({"error": "Could not read the tracker workbook.", "details": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR)
        if parsed.path == "/api/download":
            if not WORKBOOK.exists():
                return self.send_json({"error": "The tracker workbook has not been created yet."}, HTTPStatus.NOT_FOUND)
            data = WORKBOOK.read_bytes()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="Dexcom_Job_Tracker.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        if parsed.path == "/":
            self.path = "/index.html"
        return super().do_GET()

    def do_POST(self):
        if urlparse(self.path).path != "/api/run":
            return self.send_json({"error": "Not found."}, HTTPStatus.NOT_FOUND)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            request = json.loads(self.rfile.read(length) or b"{}")
        except (ValueError, json.JSONDecodeError):
            return self.send_json({"error": "Invalid JSON request."}, HTTPStatus.BAD_REQUEST)
        filter_name = request.get("filter", "7days")
        if filter_name not in {"7days", "24hours"}:
            return self.send_json({"error": "filter must be 7days or 24hours."}, HTTPStatus.BAD_REQUEST)
        with state_lock:
            if state["status"] == "running":
                return self.send_json({"error": "A tracker run is already in progress."}, HTTPStatus.CONFLICT)
            run_id = uuid.uuid4().hex
            run = {"id": run_id, "status": "running", "filter": filter_name, "started_at": datetime.now().isoformat(timespec="seconds"), "stats": None, "error": None, "details": None}
            state["status"] = "running"
            state["runs"][run_id] = run
        threading.Thread(target=launch_run, args=(run_id, filter_name), daemon=True).start()
        return self.send_json(run, HTTPStatus.ACCEPTED)


def main() -> None:
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    server = ThreadingHTTPServer((host, port), DashboardHandler)
    print(f"Dexcom Job Tracker dashboard: http://{host}:{port}")
    print("Press Ctrl+C to stop.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nDashboard stopped.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
