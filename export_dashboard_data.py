#!/usr/bin/env python3
"""Export the canonical Dexcom workbook as static dashboard data.

The Excel workbook remains the source of truth. This script only produces the
browser-readable JSON representation committed by the scheduled workflow.
"""
from __future__ import annotations

import argparse
import json
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from tracker_engine import HEADERS, SHEET_NAME, canonical_url


def jobs_from_workbook(path: Path) -> list[dict[str, str]]:
    """Read and validate the tracker workbook before exporting it."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing '{SHEET_NAME}'.")
        worksheet = workbook[SHEET_NAME]
        headers = [worksheet.cell(1, column=index).value for index in range(1, 5)]
        if headers != HEADERS:
            raise ValueError(f"Workbook columns must be exactly: {', '.join(HEADERS)}.")
        jobs: list[dict[str, str]] = []
        direct_urls: set[str] = set()
        for row_number, (country, posted_date, title, url) in enumerate(
            worksheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2
        ):
            if not any((country, posted_date, title, url)):
                continue
            canonical, job_id = canonical_url(str(url or ""))
            if not job_id or not canonical:
                raise ValueError(f"Row {row_number} has an invalid LinkedIn Direct URL.")
            if canonical in direct_urls:
                raise ValueError(f"Row {row_number} duplicates a Direct URL.")
            if not all((country, posted_date, title)):
                raise ValueError(f"Row {row_number} is missing required job data.")
            if isinstance(posted_date, datetime):
                posted = posted_date.date().isoformat()
            elif isinstance(posted_date, date):
                posted = posted_date.isoformat()
            else:
                raise ValueError(f"Row {row_number} has an invalid Job Posted Date.")
            direct_urls.add(canonical)
            jobs.append({
                "country": str(country).strip(),
                "posted_date": posted,
                "title": str(title).strip(),
                "url": canonical,
            })
        return jobs
    finally:
        workbook.close()


def export_dashboard_data(workbook_path: Path, output_path: Path) -> dict:
    jobs = jobs_from_workbook(workbook_path)
    previous: dict = {}
    if output_path.exists():
        try:
            previous = json.loads(output_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            previous = {}
    # Keep the committed update time stable when the canonical workbook data has not changed.
    generated_at = previous.get("generated_at") if previous.get("jobs") == jobs else None
    payload = {
        "generated_at": generated_at or datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "workbook": workbook_path.name,
        "jobs": jobs,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate static dashboard JSON from the tracker workbook.")
    parser.add_argument("--workbook", default="Dexcom_Job_Tracker.xlsx")
    parser.add_argument("--output", default="web/jobs.json")
    args = parser.parse_args()
    payload = export_dashboard_data(Path(args.workbook), Path(args.output))
    print(f"Exported {len(payload['jobs'])} jobs to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
