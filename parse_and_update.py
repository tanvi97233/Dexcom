import sys
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

RAW_FILE = "raw_jobs.txt"
TRACKER_FILE = "Dexcom_Job_Tracker.xlsx"
SHEET_NAME = "Dexcom Job Tracker"
HEADERS = ["Country", "Job Posted Date", "Job Title", "Direct URL"]


def parse_raw(path):
    rows, bad = [], []
    with open(path, encoding="utf-8") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            parts = [p.strip() for p in line.split("|")]
            if len(parts) != 4:
                bad.append((lineno, line, "wrong number of fields"))
                continue
            country, date_str, title, job_id = parts
            job_id = re.sub(r"\D", "", job_id)  # keep digits only
            if not job_id:
                bad.append((lineno, line, "no numeric job id"))
                continue
            try:
                date_val = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                bad.append((lineno, line, "bad date format, expected YYYY-MM-DD"))
                continue
            if not country or not title:
                bad.append((lineno, line, "missing country or title"))
                continue
            url = f"https://www.linkedin.com/jobs/view/{job_id}"
            rows.append({"country": country, "date": date_val, "title": title,
                         "url": url, "job_id": job_id})
    return rows, bad


def load_or_create_tracker(path):
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[SHEET_NAME]
        existing_ids = set()
        for r in range(2, ws.max_row + 1):
            url = ws.cell(row=r, column=4).value
            if url:
                m = re.search(r"(\d+)$", str(url))
                if m:
                    existing_ids.add(m.group(1))
        return wb, ws, existing_ids
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        for col, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(name="Arial", bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E78")
        return wb, ws, set()


def write_rows(ws, rows):
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    next_row = ws.max_row + 1
    for row in rows:
        ws.cell(row=next_row, column=1, value=row["country"]).font = Font(name="Arial", size=10)
        c2 = ws.cell(row=next_row, column=2, value=row["date"])
        c2.number_format = "yyyy-mm-dd"
        c2.font = Font(name="Arial", size=10)
        ws.cell(row=next_row, column=3, value=row["title"]).font = Font(name="Arial", size=10)
        c4 = ws.cell(row=next_row, column=4, value=row["url"])
        c4.hyperlink = row["url"]
        c4.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        for col in range(1, 5):
            ws.cell(row=next_row, column=col).border = border
        next_row += 1


def write_needs_review(wb, bad):
    if "Needs Review" in wb.sheetnames:
        del wb["Needs Review"]
    if not bad:
        return
    ws = wb.create_sheet("Needs Review")
    ws.append(["Line #", "Raw line", "Issue"])
    for lineno, line, issue in bad:
        ws.append([lineno, line, issue])


def main():
    rows, bad = parse_raw(RAW_FILE)
    wb, ws, existing_ids = load_or_create_tracker(TRACKER_FILE)
    new_rows = [r for r in rows if r["job_id"] not in existing_ids]
    skipped_dupes = len(rows) - len(new_rows)
    write_rows(ws, new_rows)
    write_needs_review(wb, bad)
    for i, w in enumerate([20, 18, 55, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(TRACKER_FILE)
    print(f"Added {len(new_rows)} new rows. Skipped {skipped_dupes} duplicates. "
          f"{len(bad)} malformed lines sent to 'Needs Review'.")


if __name__ == "__main__":
    main()
